import os
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.chart import BarChart, Reference
import re
from openpyxl.drawing.image import Image
import unicodedata


# Configurar logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def insertar_formato(ws, start_row, start_col, nrows, ncols):
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    for row in ws.iter_rows(min_row=start_row, 
                            max_row=start_row + nrows + 1, 
                            min_col=start_col, 
                            max_col=start_col + ncols):
        for cell in row:
            cell.font = Font()              
            cell.fill = PatternFill()       
            cell.border = no_border         
            cell.alignment = Alignment()    
            if cell.value is not None and isinstance(cell.value, int) and 1900 <= cell.value <= 2100:
                pass
            else:
                cell.number_format = '#,##0.00'
                
    if ws.title in ["Países productores", "Países exportadores", "Países importadores"]:
        for row in ws.iter_rows(min_row=12, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

def eliminar_graficas(ws):
    try:
        if hasattr(ws, '_charts'):
            ws._charts = []
        return True
    except Exception as e:
        logger.error(f"Error al eliminar gráficas de {ws.title}: {e}")
        return False

def extraer_producto(nombre_archivo):
    patron = r'Mercado mundial\s*-\s*(.+?)\.xlsx'
    resultado = re.search(patron, nombre_archivo, re.IGNORECASE)
    return resultado.group(1).strip() if resultado else None

def seleccionar_plantilla(directorio, template_path, template_names):
    mapping = {
        'Plantilla A': template_names[0],
        'Plantilla B': template_names[1],
        'Plantilla C': template_names[2],
        'Plantilla D': template_names[3],
        'Plantilla E': template_names[4],
        'Plantilla F': template_names[5],
        'Plantilla G': template_names[6],
        'Plantilla H': template_names[7],
        'Plantilla I': template_names[8]
    }
    for key, template in mapping.items():
        if directorio.endswith(key):
            return os.path.join(template_path, template)
    # Por defecto
    return os.path.join(template_path, template_names[0])

def crear_graficas_anuales(ws, fila_encabezado, col_anio):
    # Determinar la fila de inicio de datos
    fila_data = fila_encabezado + 1

    # Recolectar las filas en las que la celda en la columna de años NO es None
    data_rows = [r for r in range(fila_data, ws.max_row + 1) 
                 if ws.cell(row=r, column=col_anio).value is not None]
    if not data_rows:
        logger.info("No se encontraron datos para graficar")
        return False
    ultima_fila = data_rows[-1]

    # Contar las columnas con datos en la primera fila de datos
    valid_cols = 0
    for col in range(col_anio, ws.max_column + 1):
        if ws.cell(row=fila_data, column=col).value is not None:
            valid_cols += 1

    # Si solo existe la columna de años, usar solo los últimos 10 registros
    if valid_cols == 1:
        start_chart = max(fila_data, ultima_fila - 10 + 1)
    else:
        start_chart = fila_data

    # Referencias para categorías y datos
    year_categories = Reference(ws,
                                min_col=col_anio,
                                min_row=start_chart,
                                max_row=ultima_fila)
    chart_value = BarChart()
    chart_value.type = "col"
    
    # Para el rango de datos:
    # Si solo hay la columna de años, se asume que se quiere graficar esos valores (o se puede ajustar la lógica)
    if valid_cols == 1:
        data_valor = Reference(ws,
                               min_col=col_anio,
                               min_row=start_chart,
                               max_row=ultima_fila)
    else:
        data_valor = Reference(ws,
                               min_col=col_anio + 1,
                               min_row=start_chart,
                               max_row=ultima_fila)
    chart_value.add_data(data_valor, titles_from_data=False)
    chart_value.set_categories(year_categories)
    
    chart_value.x_axis.scaling.orientation = "maxMin"
    chart_value.x_axis.delete = False   
    chart_value.y_axis.delete = False   
    chart_value.y_axis.scaling.min = 0
    chart_value.legend = None
    chart_value.series[0].graphicalProperties.solidFill = "4472C4"
    
    # Calcular el valor máximo en la serie para definir el formato del eje (se asume que los datos a graficar están en data_valor)
    max_val = None
    for row in ws.iter_rows(min_row=start_chart, max_row=ultima_fila, min_col=(col_anio if valid_cols == 1 else col_anio+1), max_col=(col_anio if valid_cols == 1 else col_anio+1)):
        for cell in row:
            try:
                valor = float(cell.value)
                if max_val is None or valor > max_val:
                    max_val = valor
            except (ValueError, TypeError):
                continue

    # Formateo del eje Y y título del gráfico
    if max_val is not None:
        if max_val >= 1_000_000:
            numeral_label = "Millones"
            chart_value.y_axis.number_format = '#,##0.00,,'
        elif max_val >= 1000:
            numeral_label = "Miles"
            chart_value.y_axis.number_format = '#,##0.00,'
        else:
            numeral_label = ""
            chart_value.y_axis.number_format = '#,##0.00'
        
        cell_c10 = ws["C10"].value if ws["C10"].value is not None else ""
        cell_c11 = ws["C11"].value if ws["C11"].value is not None else ""
        cell_c11_clean = str(cell_c11).strip("()")
        
        if numeral_label:
            chart_value.title = f"{cell_c10} ({numeral_label} de {cell_c11_clean})"
        else:
            chart_value.title = cell_c10
    else:
        chart_value.title = ws["C10"].value if ws["C10"].value is not None else ""
    chart_value.title.overlay = False

    ws.add_chart(chart_value, "I10")
    return True

# Configuración de rutas
source_dir = r"../Datos_Extraidos"
template_path = r"../estadisticas_macro_shared/estadisticas_macro_shared/Plantillas"
output_dir = r"Resultados"
template_names = [n for n in os.listdir(template_path) if n.endswith('.xlsx') and n.startswith('Mercado mundial')]

if not os.path.exists(template_path):
    raise FileNotFoundError(f"La plantilla no fue encontrada en: {template_path}")

os.makedirs(output_dir, exist_ok=True)

# Recorremos cada directorio en source_dir (por ejemplo, Plantilla A, B o C)
for directorio in os.listdir(source_dir):
    dir_path = os.path.join(source_dir, directorio)
    if not os.path.isdir(dir_path):
        continue

    files = [f for f in os.listdir(dir_path) if f.endswith('.xlsx')]
    for file in files:
        producto = extraer_producto(file)
        source_file = os.path.join(dir_path, file)
        datos_archivo = pd.read_excel(source_file, sheet_name=None)
        base_name = os.path.splitext(file)[0]
        output_file = os.path.join(output_dir, f"{base_name}.xlsx")
        
        workbook_path = seleccionar_plantilla(os.path.basename(dir_path), template_path, template_names)

        
        book = load_workbook(workbook_path)
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        writer._book = book
        writer._sheets = {ws.title: ws for ws in book.worksheets}
        
        logger.info(f"Plantilla utilizada: {workbook_path}")
        logger.info(f"Hojas en plantilla: {book.sheetnames}")
        
        for sheet in list(book.sheetnames):
            if sheet != "(Paises)" and sheet in datos_archivo:
                df = datos_archivo[sheet]
                df.to_excel(writer, sheet_name=sheet, index=False, startrow=11, startcol=1, header=False)
                ws = book[sheet]
                ws.cell(row=6, column=3, value=producto)
                ws.sheet_view.showGridLines = False
                nrows, ncols = df.shape
                insertar_formato(ws, start_row=12, start_col=2, nrows=nrows, ncols=ncols)
                eliminar_graficas(ws)
                crear_graficas_anuales(ws, fila_encabezado=11, col_anio=2)

        
        
        if "(Paises)" in book.sheetnames:
            for sheet in datos_archivo.keys():
                if sheet not in book.sheetnames:
                    new_sheet = book.copy_worksheet(book["(Paises)"])
                    new_sheet.title = sheet  
                    img_path = 'fira.png'
                    imagen_fira = Image(img_path)
                    imagen_fira.anchor = 'J2'
                    new_sheet.add_image(imagen_fira)
                    writer._sheets[new_sheet.title] = new_sheet
                    df = datos_archivo[sheet]
                    df.columns = df.columns.astype(str).str.replace(r'^Unnamed.*$', '', regex=True)
                    df.to_excel(writer, sheet_name=new_sheet.title, index=False, startrow=11, startcol=1, header=False)
                    new_sheet.cell(row=6, column=3, value=producto)
                    new_sheet.sheet_view.showGridLines = False
                    insertar_formato(new_sheet, start_row=12, start_col=2, nrows=df.shape[0], ncols=df.shape[1])
                    eliminar_graficas(new_sheet)
                    crear_graficas_anuales(new_sheet, fila_encabezado=11, col_anio=2)
            book.remove(book["(Paises)"])
        
        writer.close()
        logger.info(f"Procesado: {output_file}")